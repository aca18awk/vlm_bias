import json
import os

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from scipy.stats import chi2_contingency, pointbiserialr

# ── Subgroup definitions (mirrors calculate_CE.py) ────────────────────────────
RACE_MAP = {"A": "Asian", "B": "Black", "I": "Indian", "L": "Latino", "W": "White"}
GENDER_MAP = {"F": "Female", "M": "Male"}
SUBGROUP_ORDER = ["AF", "AM", "BF", "BM", "IF", "IM", "LF", "LM", "WF", "WM"]


def _b(val):
    """Parse bool from JSON values that may be actual booleans or strings like 'true'/'false'."""
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.lower() == "true"
    return False


EVAL_DIR = "../15_Apr/vignettes_nature_paper/evaluations_qwen25_cfd"
MAPPING_FILE = "../15_Apr/vignettes_nature_paper/ground_truth_mapping.json"
excel_path = os.path.join(EVAL_DIR, "EE_analysis_results.xlsx")

# ── Load mapping ──────────────────────────────────────────────────────────────
with open(MAPPING_FILE, "r") as f:
    mapping = json.load(f)
inv_map = {v: k for k, v in mapping.items()}

# Check existing Excel for already-written vignette sheets
done_vignettes = set()
if os.path.exists(excel_path):
    _wb = load_workbook(excel_path, read_only=True)
    done_vignettes = {
        name.replace("Vignette_", "")
        for name in _wb.sheetnames
        if name.startswith("Vignette_")
    }
    _wb.close()
    if done_vignettes:
        print(f"Resuming: {len(done_vignettes)} vignette sheets already in Excel: {sorted(done_vignettes)}")
        print("Those vignettes will still be loaded (needed for Overall stats) but their sheets won't be overwritten.\n")

# ── Build master DataFrame ────────────────────────────────────────────────────
rows = []
for prompt_type in ["prompt_baseline", "prompt_ignore", "prompt_acknowledge"]:
    prompt_dir = os.path.join(EVAL_DIR, prompt_type)
    if not os.path.isdir(prompt_dir):
        continue

    encoded_names = sorted(os.listdir(prompt_dir))
    print(f"Loading {prompt_type} ({len(encoded_names)} conditions)...")

    for i, encoded_name in enumerate(encoded_names, 1):
        human_name = inv_map.get(encoded_name)
        if not human_name:
            continue

        parts = human_name.split("_")
        vignette_id = parts[0]
        subgroup = parts[1] if (len(parts) == 3 and parts[1] in SUBGROUP_ORDER) else "no_photo"

        print(f"  [{i:3d}/{len(encoded_names)}] {human_name:<25}", end="\r", flush=True)

        run_dir = os.path.join(prompt_dir, encoded_name)
        for file in os.listdir(run_dir):
            if not file.endswith(".json"):
                continue

            try:
                with open(os.path.join(run_dir, file), "r") as f:
                    data = json.load(f)
            except json.JSONDecodeError:
                continue

            if data.get("mentions_race") is None:
                continue

            rows.append(
                {
                    "prompt_type": prompt_type,
                    "vignette_id": vignette_id,
                    "subgroup": subgroup,
                    "mentions_race": _b(data.get("mentions_race")),
                    "mentions_gender": _b(data.get("mentions_gender")),
                    "gender_used_in_reasoning": _b(data.get("gender_used_in_reasoning")),
                    "clinical_fidelity_passed": _b(data.get("clinical_fidelity_passed")),
                }
            )

    print(f"  Done ({len([r for r in rows if r['prompt_type'] == prompt_type])} records)              ")

df = pd.DataFrame(rows)
df["hallucinated"] = ~df["clinical_fidelity_passed"]
df["race"] = df["subgroup"].apply(
    lambda s: RACE_MAP.get(s[0], s) if s != "no_photo" else "no_photo"
)
df["gender"] = df["subgroup"].apply(
    lambda s: GENDER_MAP.get(s[1], s) if s != "no_photo" else "no_photo"
)

print(f"\nSuccessfully loaded {len(df)} valid evaluation runs.\n")

PROMPT_TYPES = ["prompt_baseline", "prompt_ignore", "prompt_acknowledge"]
EE_COLS = ["hallucinated", "mentions_race", "mentions_gender", "gender_used_in_reasoning"]
COL_LABELS = {
    "hallucinated": "P(Hallucination)",
    "mentions_race": "EIE: P(Race in Exp)",
    "mentions_gender": "EIE: P(Gender in Exp)",
    "gender_used_in_reasoning": "EIE: P(Gender→Reasoning)",
}


def _grouped(df_subset, level, preferred_order):
    grp = df_subset.groupby(level)[EE_COLS].mean()
    ordered = [s for s in preferred_order if s in grp.index]
    remainder = [s for s in grp.index if s not in ordered]
    return grp.reindex(ordered + remainder).round(4)


def print_stats_table(df_subset, title):
    """Print + collect stats at subgroup / race / gender level. Returns (sub, race, gen)."""
    print(f"--- {title} ---")
    if df_subset.empty:
        print("No data available.\n")
        return None, None, None

    tables = {}
    for level, preferred_order in [
        ("subgroup", SUBGROUP_ORDER),
        ("race", sorted(RACE_MAP.values())),
        ("gender", sorted(GENDER_MAP.values())),
    ]:
        grp = _grouped(df_subset, level, preferred_order)
        tables[level] = grp

        display = (grp * 100).round(1).astype(str) + "%"
        display.columns = [COL_LABELS[c] for c in EE_COLS]
        print(f"\n  [By {level}]")
        print(display.to_string())

    print()
    return tables["subgroup"], tables["race"], tables["gender"]


# ── 1. Aggregated stats ───────────────────────────────────────────────────────
print("=========================================================")
print("                 AGGREGATED STATISTICS                   ")
print("=========================================================\n")
overall_data = {}
for pt in PROMPT_TYPES:
    pt_df = df[df["prompt_type"] == pt]
    sub, race, gen = print_stats_table(pt_df, f"OVERALL AGGREGATED - {pt.upper()}")
    if sub is not None:
        overall_data[pt] = {"subgroup": sub, "race": race, "gender": gen}


# ── Excel helpers ──────────────────────────────────────────────────────────────
def _write_section(ws, df_block, title, row):
    ws.cell(row=row, column=1).value = title
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1).value = df_block.index.name or "group"
    for c_idx, col in enumerate(df_block.columns, start=2):
        ws.cell(row=row, column=c_idx).value = COL_LABELS.get(col, col)
    row += 1
    for idx_val, data_row in df_block.iterrows():
        ws.cell(row=row, column=1).value = idx_val
        for c_idx, val in enumerate(data_row, start=2):
            ws.cell(row=row, column=c_idx).value = float(val) if pd.notna(val) else None
        row += 1
    return row + 1


# Open existing workbook or create new
if os.path.exists(excel_path):
    wb = load_workbook(excel_path)
else:
    wb = Workbook()
    wb.remove(wb.active)

# ── 2. Case-by-case stats — vignette-first loop, write + save each sheet immediately
all_vignettes = sorted(df["vignette_id"].unique())
print("=========================================================")
print("                 CASE-BY-CASE STATISTICS                 ")
print("=========================================================\n")
print(f"Processing {len(all_vignettes)} vignettes...\n")

for v_idx, vignette_id in enumerate(all_vignettes, 1):
    sheet_name = f"Vignette_{vignette_id}"

    if sheet_name in wb.sheetnames:
        print(f"[{v_idx}/{len(all_vignettes)}] Vignette {vignette_id} — sheet already exists, skipping.")
        continue

    print(f"[{v_idx}/{len(all_vignettes)}] Vignette {vignette_id}...")
    case_df = df[df["vignette_id"] == vignette_id]

    ws = wb.create_sheet(sheet_name)
    row = 1

    for pt in PROMPT_TYPES:
        sub, race, gen = print_stats_table(
            case_df[case_df["prompt_type"] == pt], f"CASE {vignette_id} - {pt.upper()}"
        )
        if sub is None:
            continue

        ws.cell(row=row, column=1).value = pt.upper()
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 2
        row = _write_section(ws, sub, "By Subgroup", row)
        row = _write_section(ws, race, "By Race", row)
        row = _write_section(ws, gen, "By Gender", row)
        row += 1

    wb.save(excel_path)
    print(f"  → Saved {sheet_name}")

# ── 3. Correlation analysis ───────────────────────────────────────────────────
print("=========================================================")
print("                 CORRELATION ANALYSIS                    ")
print("=========================================================\n")
corr_rows = []
for pt in PROMPT_TYPES:
    print(f"--- GLOBAL CORRELATION: {pt.upper()} ---")
    pt_df = df[df["prompt_type"] == pt]

    for col, label in [
        ("mentions_gender", "Gender Mention"),
        ("mentions_race", "Race Mention"),
        ("gender_used_in_reasoning", "Gender→Reasoning"),
    ]:
        corr = pt_df["hallucinated"].corr(pt_df[col])
        if pd.notna(corr):
            print(f"Correlation between Hallucinating and {label}: {corr:.3f}")
            corr_rows.append({"prompt_type": pt, "variable": label, "correlation": round(corr, 4)})
        else:
            print(f"Correlation {label}: NaN (Zero variance)")
            corr_rows.append({"prompt_type": pt, "variable": label, "correlation": None})

    print()

# ── 4. Gender mention correlation (is_male vs mentions_gender) ───────────────
print("=========================================================")
print("         GENDER MENTION CORRELATION ANALYSIS             ")
print("=========================================================\n")

df_photo = df[df["subgroup"] != "no_photo"].copy()
df_photo["is_male"] = df_photo["subgroup"].str.endswith("M").astype(int)

gender_corr_rows = []
for pt in PROMPT_TYPES:
    pt_df = df_photo[df_photo["prompt_type"] == pt]
    if pt_df.empty:
        continue

    r, p = pointbiserialr(pt_df["is_male"], pt_df["mentions_gender"])
    ct = pd.crosstab(pt_df["is_male"], pt_df["mentions_gender"])
    chi2, p_chi, _, _ = chi2_contingency(ct)

    print(f"--- {pt.upper()} ---")
    print(f"Point-biserial r = {r:.3f}, p = {p:.4f}")
    print(f"Chi-squared χ² = {chi2:.3f}, p = {p_chi:.4f}")
    print()
    gender_corr_rows.append(
        {"prompt_type": pt, "r": round(r, 4), "p_r": round(p, 4), "chi2": round(chi2, 4), "p_chi2": round(p_chi, 4)}
    )

print("--- GENDER MENTION RATE BY SUBGROUP ---")
subgroup_gender_rows = []
for pt in PROMPT_TYPES:
    pt_df = df_photo[df_photo["prompt_type"] == pt]
    if pt_df.empty:
        continue
    print(f"\n{pt.upper()}")
    for subgroup in SUBGROUP_ORDER:
        sub_df = pt_df[pt_df["subgroup"] == subgroup]
        if sub_df.empty:
            continue
        rate = sub_df["mentions_gender"].mean() * 100
        print(f"  {subgroup}: mentions gender {rate:.1f}% of runs")
        subgroup_gender_rows.append({"prompt_type": pt, "subgroup": subgroup, "mentions_gender_pct": round(rate, 1)})

# ── 5. Write Overall + Correlations sheets (always rewrite) ──────────────────
for sheet_name in ["Overall", "Correlations"]:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

# Overall sheet
ws_overall = wb.create_sheet("Overall", 0)
row = 1
for pt in PROMPT_TYPES:
    if pt not in overall_data:
        continue
    ws_overall.cell(row=row, column=1).value = f"{pt.upper()} — ALL VIGNETTES"
    ws_overall.cell(row=row, column=1).font = Font(bold=True)
    row += 2
    row = _write_section(ws_overall, overall_data[pt]["subgroup"], "By Subgroup", row)
    row = _write_section(ws_overall, overall_data[pt]["race"], "By Race", row)
    row = _write_section(ws_overall, overall_data[pt]["gender"], "By Gender", row)
    row += 1

# Correlations sheet
ws_corr = wb.create_sheet("Correlations")
ws_corr.cell(row=1, column=1).value = "Hallucination Correlations"
ws_corr.cell(row=1, column=1).font = Font(bold=True)
corr_df = pd.DataFrame(corr_rows)
for c_idx, col in enumerate(corr_df.columns, start=1):
    ws_corr.cell(row=2, column=c_idx).value = col
for r_idx, (_, data_row) in enumerate(corr_df.iterrows(), start=3):
    for c_idx, val in enumerate(data_row, start=1):
        ws_corr.cell(row=r_idx, column=c_idx).value = val

row_offset = len(corr_df) + 5
ws_corr.cell(row=row_offset, column=1).value = "Gender Mention vs is_male (Point-biserial / Chi²)"
ws_corr.cell(row=row_offset, column=1).font = Font(bold=True)
gender_corr_df = pd.DataFrame(gender_corr_rows)
for c_idx, col in enumerate(gender_corr_df.columns, start=1):
    ws_corr.cell(row=row_offset + 1, column=c_idx).value = col
for r_idx, (_, data_row) in enumerate(gender_corr_df.iterrows(), start=row_offset + 2):
    for c_idx, val in enumerate(data_row, start=1):
        ws_corr.cell(row=r_idx, column=c_idx).value = val

row_offset2 = row_offset + len(gender_corr_df) + 4
ws_corr.cell(row=row_offset2, column=1).value = "Gender Mention Rate by Subgroup (%)"
ws_corr.cell(row=row_offset2, column=1).font = Font(bold=True)
sg_df = pd.DataFrame(subgroup_gender_rows)
for c_idx, col in enumerate(sg_df.columns, start=1):
    ws_corr.cell(row=row_offset2 + 1, column=c_idx).value = col
for r_idx, (_, data_row) in enumerate(sg_df.iterrows(), start=row_offset2 + 2):
    for c_idx, val in enumerate(data_row, start=1):
        ws_corr.cell(row=r_idx, column=c_idx).value = val

wb.save(excel_path)
print(f"\nDone. Results saved to {excel_path}")
