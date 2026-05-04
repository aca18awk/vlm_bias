import glob
import json
import os
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

RESULTS_DIR = "../15_Apr/vignettes_nature_paper/results_llama32_cfd"
MAPPING_FILE = "../15_Apr/vignettes_nature_paper/ground_truth_mapping.json"
CFD_DIR = "../images/CFD"
PROMPT_TYPES = ["prompt_baseline", "prompt_ignore", "prompt_acknowledge"]
excel_path = os.path.join(RESULTS_DIR, "photo_admit_rates.xlsx")

SUBGROUP_ORDER = ["AF", "AM", "BF", "BM", "IF", "IM", "LF", "LM", "WF", "WM"]

with open(MAPPING_FILE) as f:
    mapping = json.load(f)
inv_map = {v: k for k, v in mapping.items()}

# ── Build photo_id -> CFD filename map ────────────────────────────────────────
cfd_by_subgroup = defaultdict(list)
for path in sorted(glob.glob(os.path.join(CFD_DIR, "**", "CFD-*-N.jpg"), recursive=True)):
    fname = os.path.basename(path)
    parts = fname.split("-")
    if len(parts) >= 2:
        cfd_by_subgroup[parts[1]].append(fname.replace(".jpg", ""))

PHOTO_CFD_NAME = {}
for sg, names in cfd_by_subgroup.items():
    for idx, cfd_name in enumerate(names, 1):
        PHOTO_CFD_NAME[f"{sg}_{idx}"] = cfd_name

# ── Load data ─────────────────────────────────────────────────────────────────
rows = []
for prompt_type in PROMPT_TYPES:
    prompt_dir = os.path.join(RESULTS_DIR, prompt_type)
    if not os.path.isdir(prompt_dir):
        continue

    encoded_names = sorted(os.listdir(prompt_dir))
    print(f"Loading {prompt_type} ({len(encoded_names)} conditions)...")

    for encoded_name in encoded_names:
        human_name = inv_map.get(encoded_name)
        if not human_name:
            continue

        parts = human_name.split("_")
        if len(parts) != 3:
            continue  # skip no_photo

        vignette_id, subgroup, run_num = parts
        if subgroup not in SUBGROUP_ORDER:
            continue

        photo_id = f"{subgroup}_{run_num}"

        run_dir = os.path.join(prompt_dir, encoded_name)
        admit_count = 0
        total_valid = 0

        for file in os.listdir(run_dir):
            if not file.endswith(".json"):
                continue
            try:
                with open(os.path.join(run_dir, file)) as f:
                    data = json.load(f)
            except json.JSONDecodeError:
                continue
            ans = data.get("parsed_answer", "")
            if ans in ["admit", "discharge"]:
                if ans == "admit":
                    admit_count += 1
                total_valid += 1

        if total_valid > 0:
            rows.append({
                "prompt_type": prompt_type,
                "vignette_id": vignette_id,
                "subgroup": subgroup,
                "run_num": int(run_num),
                "photo_id": photo_id,
                "p_admit": admit_count / total_valid,
                "total_runs": total_valid,
            })

df = pd.DataFrame(rows)
print(f"\nLoaded {len(df)} condition records.\n")

# ── Helpers ───────────────────────────────────────────────────────────────────
BOLD = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TOP_FILL = PatternFill("solid", fgColor="C6EFCE")    # green — highest
BOT_FILL = PatternFill("solid", fgColor="FFC7CE")    # red — lowest
N_HIGHLIGHT = 5  # how many top/bottom photos to colour


def _write_photo_table(ws, table_df, start_row, title, cols):
    """Write a ranked photo table; returns next free row."""
    ws.cell(row=start_row, column=1).value = title
    ws.cell(row=start_row, column=1).font = BOLD
    start_row += 1

    for c_idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=start_row, column=c_idx)
        cell.value = col
        cell.font = BOLD
        cell.fill = HEADER_FILL
    start_row += 1

    n = len(table_df)
    for r_idx, (_, row) in enumerate(table_df.iterrows()):
        fill = None
        if r_idx < N_HIGHLIGHT:
            fill = TOP_FILL
        elif r_idx >= n - N_HIGHLIGHT:
            fill = BOT_FILL

        for c_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=start_row, column=c_idx)
            val = row[col]
            cell.value = float(val) if isinstance(val, float) else val
            if fill:
                cell.fill = fill
        start_row += 1

    return start_row + 1


# ── Build overall summary ─────────────────────────────────────────────────────
def overall_photo_table(pt):
    pt_df = df[df["prompt_type"] == pt]
    grp = pt_df.groupby("photo_id")["p_admit"].agg(
        avg_admit_rate="mean", std="std", n_vignettes="count"
    ).round(3).reset_index()
    grp["cfd_name"] = grp["photo_id"].map(PHOTO_CFD_NAME)
    grp = grp[["photo_id", "cfd_name", "avg_admit_rate", "std", "n_vignettes"]]
    grp = grp.sort_values("avg_admit_rate", ascending=False).reset_index(drop=True)
    return grp


# ── Build per-vignette photo table ────────────────────────────────────────────
def vignette_photo_table(vignette_id, pt):
    sub = df[(df["vignette_id"] == vignette_id) & (df["prompt_type"] == pt)].copy()
    sub = sub.sort_values("p_admit", ascending=False)[["photo_id", "p_admit", "total_runs"]]
    sub = sub.rename(columns={"p_admit": "admit_rate", "total_runs": "n_runs"})
    return sub.reset_index(drop=True)


# ── Print summary ─────────────────────────────────────────────────────────────
for pt in PROMPT_TYPES:
    tbl = overall_photo_table(pt)
    print(f"=== {pt.upper()} — TOP 5 ===")
    print(tbl.head(5).to_string(index=False))
    print(f"=== BOTTOM 5 ===")
    print(tbl.tail(5).to_string(index=False))
    print()

# ── Write Excel ───────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# Overall sheet
ws_overall = wb.create_sheet("Overall")
row = 1
ws_overall.cell(row=row, column=1).value = "Average admit rate per photo — averaged across all 12 vignettes"
ws_overall.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 1
ws_overall.cell(row=row, column=1).value = (
    f"Green = top {N_HIGHLIGHT} highest admit rate   |   Red = bottom {N_HIGHLIGHT} lowest admit rate"
)
row += 2

OVR_COLS = ["photo_id", "cfd_name", "avg_admit_rate", "std", "n_vignettes"]
for pt in PROMPT_TYPES:
    tbl = overall_photo_table(pt)
    row = _write_photo_table(ws_overall, tbl, row, pt.upper(), OVR_COLS)

col_widths = [10, 28, 16, 8, 14]
for col_idx, w in enumerate(col_widths, start=1):
    ws_overall.column_dimensions[get_column_letter(col_idx)].width = w

# Per-vignette sheets
all_vignettes = sorted(df["vignette_id"].unique())
VIG_COLS = ["photo_id", "admit_rate", "n_runs"]

for vignette_id in all_vignettes:
    ws_v = wb.create_sheet(f"Vignette_{vignette_id}")
    row = 1
    ws_v.cell(row=row, column=1).value = f"Vignette {vignette_id} — admit rate per photo"
    ws_v.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 2

    for pt in PROMPT_TYPES:
        tbl = vignette_photo_table(vignette_id, pt)
        if tbl.empty:
            continue
        row = _write_photo_table(ws_v, tbl, row, pt.upper(), VIG_COLS)

    for col_idx in range(1, len(VIG_COLS) + 1):
        ws_v.column_dimensions[get_column_letter(col_idx)].width = 16

wb.save(excel_path)
print(f"Done. Saved to {excel_path}")
