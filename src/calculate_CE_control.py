"""
Control analysis: grey rectangle vs. no-photo vs. demographic photo.

The control condition replaces the patient photo with a plain grey rectangle.
This isolates the effect of *any* image (image presence) from the effect of
*demographic* information in the image (image content).

Comparisons computed per vignette × prompt type:
  CE(grey vs none)  = grey_rect admit rate  − no_photo admit rate
  CE(demo vs none)  = cfd_mean admit rate   − no_photo admit rate  (reference)
  CE(grey vs demo)  = grey_rect admit rate  − cfd_mean admit rate
"""

import json
import os

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# ── Paths ─────────────────────────────────────────────────────────────────────
CONTROL_DIR = "../15_Apr/vignettes_nature_paper/results_llama32_control"
CFD_DIR     = "../15_Apr/vignettes_nature_paper/results_llama32_cfd"

CONTROL_MAPPING = "../15_Apr/vignettes_nature_paper/ground_truth_mapping_control.json"
CFD_MAPPING     = "../15_Apr/vignettes_nature_paper/ground_truth_mapping.json"

SUBGROUP_ORDER = ["AF", "AM", "BF", "BM", "IF", "IM", "LF", "LM", "WF", "WM"]
PROMPT_TYPES   = ["prompt_baseline", "prompt_ignore", "prompt_acknowledge"]

excel_path = os.path.join(CONTROL_DIR, "control_analysis.xlsx")

# ── Load mappings ─────────────────────────────────────────────────────────────
with open(CONTROL_MAPPING, "r") as f:
    control_mapping = json.load(f)   # {"01_grey_rect": "01_816", ...}

with open(CFD_MAPPING, "r") as f:
    cfd_mapping = json.load(f)       # {"01_AF_1": "...", "01_no_photo": "...", ...}

control_inv = {v: k for k, v in control_mapping.items()}
cfd_inv     = {v: k for k, v in cfd_mapping.items()}


# ── Helper: load admit rate from a result folder ───────────────────────────────
def load_admit_rate(result_dir):
    admit, total = 0, 0
    for fname in os.listdir(result_dir):
        if not fname.endswith(".json"):
            continue
        with open(os.path.join(result_dir, fname), "r") as f:
            data = json.load(f)
        ans = data.get("parsed_answer", "")
        if ans in ("admit", "discharge"):
            if ans == "admit":
                admit += 1
            total += 1
    return (admit / total, total) if total > 0 else (None, 0)


# ── 1. Load grey-rectangle results ────────────────────────────────────────────
print("Loading control (grey rectangle) results...")
grey_rows = []
for prompt_type in PROMPT_TYPES:
    prompt_dir = os.path.join(CONTROL_DIR, prompt_type)
    if not os.path.isdir(prompt_dir):
        continue
    encoded_names = sorted(os.listdir(prompt_dir))
    print(f"  {prompt_type} ({len(encoded_names)} conditions)...", end=" ")
    for encoded_name in encoded_names:
        human_name = control_inv.get(encoded_name)
        if not human_name:
            continue
        vignette_id = human_name.split("_")[0]
        p_admit, n = load_admit_rate(os.path.join(prompt_dir, encoded_name))
        if p_admit is not None:
            grey_rows.append({
                "prompt_type": prompt_type,
                "vignette_id": vignette_id,
                "grey_admit_rate": round(p_admit, 4),
                "grey_n": n,
            })
    print("done")

grey_df = pd.DataFrame(grey_rows)

# ── 2. Load CFD results: no_photo + demographic subgroup means ────────────────
print("\nLoading CFD (no_photo + demographic) results...")
cfd_rows = []
for prompt_type in PROMPT_TYPES:
    prompt_dir = os.path.join(CFD_DIR, prompt_type)
    if not os.path.isdir(prompt_dir):
        continue
    encoded_names = sorted(os.listdir(prompt_dir))
    print(f"  {prompt_type} ({len(encoded_names)} conditions)...")
    for i, encoded_name in enumerate(encoded_names, 1):
        if i % 20 == 0 or i == len(encoded_names):
            print(f"    [{i:3d}/{len(encoded_names)}]", end="\r", flush=True)
        human_name = cfd_inv.get(encoded_name)
        if not human_name:
            continue
        parts = human_name.split("_")
        vignette_id = parts[0]
        subgroup = parts[1] if (len(parts) == 3 and parts[1] in SUBGROUP_ORDER) else "no_photo"
        p_admit, n = load_admit_rate(os.path.join(prompt_dir, encoded_name))
        if p_admit is not None:
            cfd_rows.append({
                "prompt_type": prompt_type,
                "vignette_id": vignette_id,
                "subgroup": subgroup,
                "p_admit": round(p_admit, 4),
                "n": n,
            })
    print(f"  {prompt_type} — done ({len(encoded_names)} conditions)    ")

cfd_df = pd.DataFrame(cfd_rows)

# ── 3. Build comparison table ─────────────────────────────────────────────────
print("\nBuilding comparison table...")

# Restrict to vignettes present in the control dataset
control_vignettes = set(grey_df["vignette_id"].unique())

no_photo_df = (
    cfd_df[cfd_df["subgroup"] == "no_photo"]
    .rename(columns={"p_admit": "no_photo_admit_rate", "n": "no_photo_n"})
    [["prompt_type", "vignette_id", "no_photo_admit_rate", "no_photo_n"]]
)

demo_mean_df = (
    cfd_df[
        (cfd_df["subgroup"] != "no_photo") &
        (cfd_df["vignette_id"].isin(control_vignettes))
    ]
    .groupby(["prompt_type", "vignette_id"])["p_admit"]
    .mean()
    .round(4)
    .reset_index()
    .rename(columns={"p_admit": "cfd_demo_mean_admit_rate"})
)

comp = (
    grey_df
    .merge(no_photo_df, on=["prompt_type", "vignette_id"], how="left")
    .merge(demo_mean_df, on=["prompt_type", "vignette_id"], how="left")
)

comp["ce_grey_vs_none"] = (comp["grey_admit_rate"] - comp["no_photo_admit_rate"]).round(4)
comp["ce_demo_vs_none"] = (comp["cfd_demo_mean_admit_rate"] - comp["no_photo_admit_rate"]).round(4)
comp["ce_grey_vs_demo"] = (comp["grey_admit_rate"] - comp["cfd_demo_mean_admit_rate"]).round(4)

# ── 4. Print results ──────────────────────────────────────────────────────────
print(f"\n{'='*70}")
print("  CONTROL ANALYSIS: Grey Rectangle vs. No-Photo vs. CFD Demo Mean")
print(f"{'='*70}")
print("""
Columns:
  grey_admit_rate       — admit rate with grey rectangle image
  no_photo_admit_rate   — admit rate with no photo (baseline)
  cfd_demo_mean         — mean admit rate across all 10 CFD subgroups
  CE(grey−none)         — effect of any image presence
  CE(demo−none)         — effect of demographic image (reference)
  CE(grey−demo)         — grey rect vs real demographic photo
""")

DISPLAY_COLS = [
    "vignette_id", "grey_admit_rate", "no_photo_admit_rate",
    "cfd_demo_mean_admit_rate", "ce_grey_vs_none", "ce_demo_vs_none", "ce_grey_vs_demo",
]

for pt in PROMPT_TYPES:
    pt_df = comp[comp["prompt_type"] == pt][DISPLAY_COLS].sort_values("vignette_id")
    if pt_df.empty:
        continue

    print(f"\n{'─'*70}")
    print(f"  {pt.upper()}")
    print(f"{'─'*70}")
    print(pt_df.to_string(index=False))

    means = pt_df[["grey_admit_rate", "no_photo_admit_rate", "cfd_demo_mean_admit_rate",
                   "ce_grey_vs_none", "ce_demo_vs_none", "ce_grey_vs_demo"]].mean().round(4)
    print(f"\n  MEAN across vignettes:")
    for col, val in means.items():
        print(f"    {col:<35} {val:+.4f}")

# ── 5. Export to Excel ────────────────────────────────────────────────────────
if os.path.exists(excel_path):
    wb = load_workbook(excel_path)
else:
    wb = Workbook()
    wb.remove(wb.active)

# Overall comparison sheet
for sheet_name in ["Control_Overview", "By_Vignette"]:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

# Control_Overview: summary means per prompt_type
ws_ov = wb.create_sheet("Control_Overview", 0)
ws_ov.cell(row=1, column=1).value = "Control Analysis — Mean Admit Rates & Causal Effects"
ws_ov.cell(row=1, column=1).font = Font(bold=True)

summary_cols = [
    "grey_admit_rate", "no_photo_admit_rate", "cfd_demo_mean_admit_rate",
    "ce_grey_vs_none", "ce_demo_vs_none", "ce_grey_vs_demo",
]
headers = [
    "grey_admit_rate", "no_photo_admit_rate", "cfd_demo_mean",
    "CE(grey−none)", "CE(demo−none)", "CE(grey−demo)",
]

row = 3
for pt in PROMPT_TYPES:
    pt_df = comp[comp["prompt_type"] == pt]
    if pt_df.empty:
        continue

    ws_ov.cell(row=row, column=1).value = pt.upper()
    ws_ov.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    ws_ov.cell(row=row, column=1).value = "vignette_id"
    for c_idx, h in enumerate(headers, start=2):
        ws_ov.cell(row=row, column=c_idx).value = h
    row += 1

    for _, r in pt_df[["vignette_id"] + summary_cols].sort_values("vignette_id").iterrows():
        ws_ov.cell(row=row, column=1).value = r["vignette_id"]
        for c_idx, col in enumerate(summary_cols, start=2):
            v = r[col]
            ws_ov.cell(row=row, column=c_idx).value = float(v) if pd.notna(v) else None
        row += 1

    # Mean row
    means = pt_df[summary_cols].mean().round(4)
    ws_ov.cell(row=row, column=1).value = "MEAN"
    ws_ov.cell(row=row, column=1).font = Font(bold=True)
    for c_idx, col in enumerate(summary_cols, start=2):
        ws_ov.cell(row=row, column=c_idx).value = float(means[col])
    row += 3

# By_Vignette: per-vignette sheet with all 3 prompt types side by side
all_vignettes = sorted(comp["vignette_id"].unique())
print(f"\nWriting Excel: {len(all_vignettes)} vignette sheets...")

for v_idx, vignette_id in enumerate(all_vignettes, 1):
    sheet_name = f"Vignette_{vignette_id}"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1).value = f"Vignette {vignette_id} — Control vs CFD Comparison"
    ws.cell(row=1, column=1).font = Font(bold=True)
    row = 3

    for pt in PROMPT_TYPES:
        r = comp[(comp["prompt_type"] == pt) & (comp["vignette_id"] == vignette_id)]
        if r.empty:
            continue

        ws.cell(row=row, column=1).value = pt.upper()
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for col, label in [
            ("grey_admit_rate",          "Grey rect admit rate"),
            ("no_photo_admit_rate",       "No-photo admit rate"),
            ("cfd_demo_mean_admit_rate",  "CFD demo mean admit rate"),
            ("ce_grey_vs_none",          "CE: grey − none"),
            ("ce_demo_vs_none",          "CE: demo − none"),
            ("ce_grey_vs_demo",          "CE: grey − demo"),
            ("grey_n",                   "Grey rect N runs"),
            ("no_photo_n",               "No-photo N runs"),
        ]:
            v = r.iloc[0][col] if col in r.columns else None
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=2).value = float(v) if pd.notna(v) else None
            row += 1
        row += 1

    wb.save(excel_path)
    print(f"  [{v_idx}/{len(all_vignettes)}] → Saved Vignette_{vignette_id}")

wb.save(excel_path)
print(f"\nDone. Results saved to {excel_path}")
