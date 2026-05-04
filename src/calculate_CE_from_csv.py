import os

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from scipy.stats import entropy

CSV_PATH = "Master_Aggregated_Data.csv"
CONTROL_CSV_PATH = "Control_Aggregated_Data.csv"
OUTPUT_DIR = "."
SUMMARY_CSV = os.path.join(OUTPUT_DIR, "CE_summary.csv")

RACE_MAP = {"A": "Asian", "B": "Black", "I": "Indian", "L": "Latino", "W": "White"}
GENDER_MAP = {"F": "Female", "M": "Male"}
SUBGROUP_ORDER = ["AF", "AM", "BF", "BM", "IF", "IM", "LF", "LM", "WF", "WM"]
PROMPT_ORDER = ["baseline", "ignore", "acknowledge"]
EPSILON = 1e-5

# Load photo data
df_raw = pd.read_csv(CSV_PATH)
df_raw["subgroup"] = df_raw["race"] + df_raw["gender"]
df_raw["p_admit"] = df_raw["admit_rate"]
df_raw["p_discharge"] = 1.0 - df_raw["admit_rate"]

# Load control baselines (no_photo and grey_rect) keyed by (model, prompt, vignette_id)
ctrl_raw = pd.read_csv(CONTROL_CSV_PATH)
ctrl = {}
for cond in ["no_photo", "grey_rect"]:
    sub = ctrl_raw[ctrl_raw["condition"] == cond][
        ["model", "prompt", "vignette_id", "admit_rate", "n_total"]
    ].set_index(["model", "prompt", "vignette_id"])
    ctrl[cond] = sub

models = df_raw["model"].unique()
print(f"Models found: {list(models)}\n")


def _se_ce(p, n_p, q, n_q):
    """Binomial SE of (p - q)."""
    return np.sqrt(p * (1 - p) / max(n_p, 1) + q * (1 - q) / max(n_q, 1))


def _agg_group(grp):
    """Aggregate one group (subgroup / race / gender slice) into summary stats."""
    n = len(grp)
    out = {
        "admit_mean": grp["p_admit"].mean(),
        "admit_std": grp["p_admit"].std(),
    }
    for cond in ["no_photo", "grey_rect"]:
        out[f"baseline_{cond}"] = grp[f"q_{cond}"].mean()
        out[f"ce_{cond}"] = grp[f"ce_{cond}"].mean()
        # SE of the mean CE: sqrt(mean(var_i) / n)
        out[f"se_{cond}"] = np.sqrt(grp[f"se_{cond}"].pow(2).mean() / n) if n > 0 else np.nan
    return pd.Series(out)


def print_grouped_stats(subset, label):
    """Print stats by subgroup, race, gender. Returns (sub, race, gen) DataFrames."""
    print(f"\n{'='*60}")
    print(f"  {label}")
    print(f"{'='*60}")

    def agg(groupby_obj):
        return groupby_obj.apply(_agg_group).round(4)

    print("\n  [By subgroup]")
    sub = agg(subset.groupby("subgroup"))
    sub = sub.reindex([s for s in SUBGROUP_ORDER if s in sub.index])
    print(sub.to_string())

    tmp = subset.copy()
    tmp["race_label"] = tmp["subgroup"].apply(lambda s: RACE_MAP.get(s[0], s))
    tmp["gender_label"] = tmp["subgroup"].apply(lambda s: GENDER_MAP.get(s[1], s))

    print("\n  [By race]")
    race = agg(tmp.groupby("race_label"))
    race = race.reindex(sorted(race.index))
    print(race.to_string())

    print("\n  [By gender]")
    gen = agg(tmp.groupby("gender_label"))
    print(gen.to_string())

    return sub, race, gen


def _write_section(ws, df_block, title, row):
    ws.cell(row=row, column=1).value = title
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1).value = df_block.index.name or "group"
    for c_idx, col in enumerate(df_block.columns, start=2):
        ws.cell(row=row, column=c_idx).value = col
    row += 1
    for idx_val, data_row in df_block.iterrows():
        ws.cell(row=row, column=1).value = idx_val
        for c_idx, val in enumerate(data_row, start=2):
            ws.cell(row=row, column=c_idx).value = float(val) if pd.notna(val) else None
        row += 1
    return row + 1


summary_rows = []

for model_name in models:
    df = df_raw[df_raw["model"] == model_name].copy()
    print(f"\n{'='*60}")
    print(f"  Model: {model_name}  ({len(df)} rows)")
    print(f"{'='*60}\n")

    excel_path = os.path.join(OUTPUT_DIR, f"CE_{model_name}.xlsx")

    done_vignettes = set()
    if os.path.exists(excel_path):
        _wb = load_workbook(excel_path, read_only=True)
        done_vignettes = {
            name.replace("Vignette_", "")
            for name in _wb.sheetnames
            if name.startswith("Vignette_")
        }
        _wb.close()
        if done_vignettes:
            print(f"Resuming: {len(done_vignettes)} vignette sheets already in Excel.")

    # Compute CE and SE for both baselines
    results = []
    for prompt in PROMPT_ORDER:
        prompt_df = df[df["prompt"] == prompt]
        if prompt_df.empty:
            continue

        for vignette_id in prompt_df["vignette_id"].unique():
            vig_df = prompt_df[prompt_df["vignette_id"] == vignette_id]
            key = (model_name, prompt, vignette_id)

            baselines = {}
            for cond in ["no_photo", "grey_rect"]:
                if key in ctrl[cond].index:
                    r = ctrl[cond].loc[key]
                    baselines[cond] = (float(r["admit_rate"]), int(r["n_total"]))
                else:
                    baselines[cond] = (float(vig_df["p_admit"].mean()), None)

            q_nop, n_nop = baselines["no_photo"]
            q_gry, n_gry = baselines["grey_rect"]
            q_nop = max(q_nop, EPSILON)
            q_gry = max(q_gry, EPSILON)
            Q_nop = np.array([q_nop, max(1 - q_nop, EPSILON)])

            for _, photo_row in vig_df.iterrows():
                p_admit = max(float(photo_row["p_admit"]), EPSILON)
                p_discharge = max(float(photo_row["p_discharge"]), EPSILON)
                n_photo = int(photo_row["n_total"])
                P = np.array([p_admit, p_discharge])

                results.append({
                    "prompt": prompt,
                    "vignette_id": str(vignette_id),
                    "subgroup": photo_row["subgroup"],
                    "n_photo": n_photo,
                    "p_admit": round(p_admit, 4),
                    "q_no_photo": round(q_nop, 4),
                    "ce_no_photo": round(p_admit - q_nop, 4),
                    "se_no_photo": round(_se_ce(p_admit, n_photo, q_nop, n_nop or 1), 4),
                    "q_grey_rect": round(q_gry, 4),
                    "ce_grey_rect": round(p_admit - q_gry, 4),
                    "se_grey_rect": round(_se_ce(p_admit, n_photo, q_gry, n_gry or 1), 4),
                    "kl_divergence": round(float(entropy(P, Q_nop)), 4),
                })

    results_df = pd.DataFrame(results)

    if results_df.empty:
        print(f"  No data for model {model_name}, skipping.")
        continue

    # Aggregated stats
    print("\n" + "="*60)
    print(f"  AGGREGATED STATS — {model_name.upper()} — ALL VIGNETTES")
    print("="*60)
    overall_data = {}
    for prompt in PROMPT_ORDER:
        subset = results_df[results_df["prompt"] == prompt]
        if subset.empty:
            continue
        sub, race, gen = print_grouped_stats(subset, f"{prompt.upper()} — ALL VIGNETTES")
        overall_data[prompt] = {"subgroup": sub, "race": race, "gender": gen}

        for group_type, grp_df in [("race", race), ("gender", gen)]:
            for idx_val, grp_row in grp_df.iterrows():
                summary_rows.append({
                    "model": model_name,
                    "prompt": prompt,
                    "group_type": group_type,
                    "group": idx_val,
                    "admit_mean": grp_row.get("admit_mean"),
                    "baseline_no_photo": grp_row.get("baseline_no_photo"),
                    "ce_no_photo": grp_row.get("ce_no_photo"),
                    "se_ce_no_photo": grp_row.get("se_no_photo"),
                    "baseline_grey_rect": grp_row.get("baseline_grey_rect"),
                    "ce_grey_rect": grp_row.get("ce_grey_rect"),
                    "se_ce_grey_rect": grp_row.get("se_grey_rect"),
                })

    # Write Excel
    wb = load_workbook(excel_path) if os.path.exists(excel_path) else Workbook()
    if not os.path.exists(excel_path) and wb.active is not None:
        wb.remove(wb.active)

    all_vignettes = sorted(results_df["vignette_id"].unique())
    print(f"\nProcessing {len(all_vignettes)} vignettes...\n")

    for v_idx, vignette_id in enumerate(all_vignettes, 1):
        sheet_name = f"Vignette_{vignette_id}"

        if sheet_name in wb.sheetnames:
            print(f"[{v_idx}/{len(all_vignettes)}] Vignette {vignette_id} — already exists, skipping.")
            continue

        print(f"[{v_idx}/{len(all_vignettes)}] Vignette {vignette_id}...")
        ws = wb.create_sheet(sheet_name)
        row = 1

        for prompt in PROMPT_ORDER:
            vig_subset = results_df[
                (results_df["prompt"] == prompt) &
                (results_df["vignette_id"] == vignette_id)
            ]
            if vig_subset.empty:
                continue

            q_nop = vig_subset["q_no_photo"].values[0]
            q_gry = vig_subset["q_grey_rect"].values[0]
            sub, race, gen = print_grouped_stats(
                vig_subset,
                f"Vignette {vignette_id} — {prompt} (no_photo: {q_nop:.1%}, grey_rect: {q_gry:.1%})",
            )

            ws.cell(row=row, column=1).value = (
                f"{prompt.upper()} — no_photo: {q_nop:.1%}  grey_rect: {q_gry:.1%}"
            )
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 2
            row = _write_section(ws, sub, "By Subgroup", row)
            row = _write_section(ws, race, "By Race", row)
            row = _write_section(ws, gen, "By Gender", row)

            per_case = vig_subset[[
                "subgroup", "p_admit",
                "q_no_photo", "ce_no_photo", "se_no_photo",
                "q_grey_rect", "ce_grey_rect", "se_grey_rect",
            ]].copy()
            per_case["subgroup"] = pd.Categorical(
                per_case["subgroup"], categories=SUBGROUP_ORDER, ordered=True
            )
            per_case = per_case.sort_values("subgroup").set_index("subgroup")
            row = _write_section(ws, per_case, "Per-case", row)
            row += 1

        wb.save(excel_path)
        print(f"  → Saved {sheet_name}")

    # Overall sheet
    if "Overall" in wb.sheetnames:
        del wb["Overall"]
    ws_overall = wb.create_sheet("Overall", 0)
    row = 1
    for prompt in PROMPT_ORDER:
        if prompt not in overall_data:
            continue
        ws_overall.cell(row=row, column=1).value = f"{prompt.upper()} — ALL VIGNETTES"
        ws_overall.cell(row=row, column=1).font = Font(bold=True)
        row += 2
        row = _write_section(ws_overall, overall_data[prompt]["subgroup"], "By Subgroup", row)
        row = _write_section(ws_overall, overall_data[prompt]["race"], "By Race", row)
        row = _write_section(ws_overall, overall_data[prompt]["gender"], "By Gender", row)
        row += 1

    wb.save(excel_path)
    print(f"\nDone. Results saved to {excel_path}")

# Write summary CSV (long format, one row per model × prompt × group)
if summary_rows:
    summary_df = pd.DataFrame(summary_rows)
    summary_df = summary_df.round(4)
    summary_df.to_csv(SUMMARY_CSV, index=False)
    print(f"\nSummary saved to {SUMMARY_CSV}")
