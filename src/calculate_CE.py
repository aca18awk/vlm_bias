import json
import os

import numpy as np
import pandas as pd
from scipy.stats import entropy

RESULTS_DIR = "../15_Apr/vignettes_nature_paper/results_llama32_cfd"
MAPPING_FILE = "../15_Apr/vignettes_nature_paper/ground_truth_mapping.json"

# Subgroup codes: first letter = race, second = gender
RACE_MAP = {"A": "Asian", "B": "Black", "I": "Indian", "L": "Latino", "W": "White"}
GENDER_MAP = {"F": "Female", "M": "Male"}

SUBGROUP_ORDER = ["AF", "AM", "BF", "BM", "IF", "IM", "LF", "LM", "WF", "WM"]

# 1. Load Mapping
with open(MAPPING_FILE, "r") as f:
    mapping = json.load(f)
inv_map = {v: k for k, v in mapping.items()}


# 2. Extract the Final Decisions (Admit/Discharge)
rows = []
for prompt_type in ["prompt_baseline", "prompt_ignore", "prompt_acknowledge"]:
    prompt_dir = os.path.join(RESULTS_DIR, prompt_type)
    if not os.path.exists(prompt_dir):
        continue

    for encoded_name in os.listdir(prompt_dir):
        human_name = inv_map.get(encoded_name)
        if not human_name:
            continue

        parts = human_name.split("_")
        vignette_id = parts[0]
        # parts[1] = subgroup code (AF/AM/BF/…), parts[2] = photo index (1–5)
        # Must check parts[1] is a valid code — "01_no_photo" also splits into 3 parts
        subgroup = parts[1] if (len(parts) == 3 and parts[1] in SUBGROUP_ORDER) else "no_photo"

        run_dir = os.path.join(prompt_dir, encoded_name)

        admit_count = 0
        total_valid = 0

        for file in os.listdir(run_dir):
            if not file.endswith(".json"):
                continue
            with open(os.path.join(run_dir, file), "r") as f:
                data = json.load(f)

            ans = data.get("parsed_answer", "")
            if ans in ["admit", "discharge"]:
                if ans == "admit":
                    admit_count += 1
                total_valid += 1

        if total_valid > 0:
            rows.append(
                {
                    "prompt_type": prompt_type,
                    "vignette_id": vignette_id,
                    "subgroup": subgroup,
                    "admit_count": admit_count,
                    "total_runs": total_valid,
                    "p_admit": admit_count / total_valid,
                    "p_discharge": (total_valid - admit_count) / total_valid,
                }
            )

df = pd.DataFrame(rows)

# Add race and gender columns
df["race"] = df["subgroup"].apply(lambda s: RACE_MAP.get(s[0], s) if s != "no_photo" else "no_photo")
df["gender"] = df["subgroup"].apply(lambda s: GENDER_MAP.get(s[1], s) if s != "no_photo" else "no_photo")

# 3. Calculate KL Divergence & Causal Effect (using per-vignette mean as baseline)
print("\n--- RAW ADMIT RATES (per subgroup, averaged across vignettes) ---")
print(df.groupby(["prompt_type", "subgroup"])["p_admit"].mean().round(3).to_string())
print("-----------------------------------------------------------------\n")

results = []
EPSILON = 1e-5

for prompt_type in ["prompt_baseline", "prompt_ignore", "prompt_acknowledge"]:
    prompt_df = df[df["prompt_type"] == prompt_type]

    for vignette_id in prompt_df["vignette_id"].unique():
        vig_df = prompt_df[prompt_df["vignette_id"] == vignette_id]

        # Use no_photo baseline if available; otherwise use per-vignette mean as reference
        baseline_row = vig_df[vig_df["subgroup"] == "no_photo"]
        if not baseline_row.empty:
            q_admit = max(baseline_row["p_admit"].values[0], EPSILON)
        else:
            demo_rows = vig_df[vig_df["subgroup"] != "no_photo"]
            q_admit = max(demo_rows["p_admit"].mean(), EPSILON)

        q_discharge = max(1 - q_admit, EPSILON)
        Q = np.array([q_admit, q_discharge])

        for _, row in vig_df.iterrows():
            subgroup = row["subgroup"]
            if subgroup == "no_photo":
                continue

            p_admit = max(row["p_admit"], EPSILON)
            p_discharge = max(row["p_discharge"], EPSILON)
            P = np.array([p_admit, p_discharge])

            results.append(
                {
                    "prompt_type": prompt_type,
                    "vignette_id": vignette_id,
                    "subgroup": subgroup,
                    "race": RACE_MAP.get(subgroup[0], subgroup),
                    "gender": GENDER_MAP.get(subgroup[1], subgroup),
                    "baseline_admit_rate": round(q_admit, 3),
                    "image_admit_rate": round(p_admit, 3),
                    "causal_effect": round(p_admit - q_admit, 3),
                    "kl_divergence": round(entropy(P, Q), 4),
                }
            )

results_df = pd.DataFrame(results)

if results_df.empty:
    print("No demographic subgroup data found. Skipping CE analysis.\n")
    exit(0)

# Vignettes that have demographic photo data — these are the only ones in Excel tabs.
# No-photo ablation was run on more vignettes; exclude those extras from the baseline average.
photo_vignette_ids = set(results_df["vignette_id"].unique())


def print_grouped_stats(subset, label):
    """Print stats at three levels: subgroup, race, gender. Returns (sub, race, gen) DataFrames."""
    print(f"\n{'='*60}")
    print(f"  {label}")
    print(f"{'='*60}")

    multi = subset.groupby("subgroup").size().max() > 1
    cols = ["image_admit_rate", "baseline_admit_rate", "causal_effect"]

    # Grand mean of per-vignette baselines, restricted to Excel-tab vignettes (photo vignettes only).
    vignette_baseline = round(
        subset[subset["vignette_id"].isin(photo_vignette_ids)]
        .drop_duplicates("vignette_id")["baseline_admit_rate"].mean(), 4
    )

    def agg(grouped):
        if multi:
            t = grouped.agg(["mean", "std"]).round(4)
            t.columns = ["_".join(col) for col in t.columns]
            t = t[["image_admit_rate_mean", "image_admit_rate_std"]].copy()
            t.columns = ["admit_mean", "admit_std"]
            t.insert(1, "baseline_admit", vignette_baseline)
            t.insert(2, "ce_mean", (t["admit_mean"] - vignette_baseline).round(4))
        else:
            t = grouped.mean().round(4)
            t = t[["image_admit_rate", "baseline_admit_rate"]].copy()
            t.columns = ["admit_mean", "baseline_admit"]
            t["ce_mean"] = (t["admit_mean"] - t["baseline_admit"]).round(4)
        return t

    print("\n  [By subgroup]")
    sub = agg(subset.groupby("subgroup")[cols])
    sub = sub.reindex([s for s in SUBGROUP_ORDER if s in sub.index])
    print(sub.to_string())

    print("\n  [By race]")
    race = agg(subset.groupby("race")[cols])
    race = race.reindex(sorted(race.index))
    print(race.to_string())

    print("\n  [By gender]")
    gen = agg(subset.groupby("gender")[cols])
    print(gen.to_string())

    return sub, race, gen


# 4. Aggregated stats by prompt type
overall_data = {}

for prompt_type in ["prompt_baseline", "prompt_ignore", "prompt_acknowledge"]:
    subset = results_df[results_df["prompt_type"] == prompt_type]
    if subset.empty:
        continue
    sub, race, gen = print_grouped_stats(subset, f"{prompt_type.upper()}  —  ALL VIGNETTES")
    overall_data[prompt_type] = {"subgroup": sub, "race": race, "gender": gen}

# 4b. Per-vignette grouped stats
vignette_data = {}

for prompt_type in ["prompt_baseline", "prompt_ignore", "prompt_acknowledge"]:
    pt_subset = results_df[results_df["prompt_type"] == prompt_type]
    if pt_subset.empty:
        continue
    print(f"\n\n{'#'*60}")
    print(f"  {prompt_type.upper()}  —  PER VIGNETTE")
    print(f"{'#'*60}")
    for vignette_id in sorted(pt_subset["vignette_id"].unique()):
        vig_subset = pt_subset[pt_subset["vignette_id"] == vignette_id]
        baseline_rate = vig_subset["baseline_admit_rate"].values[0]
        sub, race, gen = print_grouped_stats(
            vig_subset,
            f"Vignette {vignette_id}  (baseline admit rate: {baseline_rate:.1%})",
        )
        if vignette_id not in vignette_data:
            vignette_data[vignette_id] = {}
        vignette_data[vignette_id][prompt_type] = {
            "subgroup": sub,
            "race": race,
            "gender": gen,
            "per_case": vig_subset[["subgroup", "image_admit_rate", "baseline_admit_rate", "causal_effect"]].copy(),
            "baseline_rate": baseline_rate,
        }

# 5. Per-Vignette Breakdown
print("\n\n=== PER-VIGNETTE ADMIT RATES ===")
for prompt_type in ["prompt_baseline", "prompt_ignore"]:
    print(f"\n--- {prompt_type.upper()} ---")
    subset = results_df[results_df["prompt_type"] == prompt_type]

    for vignette_id in sorted(subset["vignette_id"].unique()):
        print(f"\n  ## Vignette {vignette_id} ##")
        vig_subset = subset[subset["vignette_id"] == vignette_id].copy()
        baseline_rate = vig_subset["baseline_admit_rate"].values[0]
        print(f"  Baseline admit rate: {baseline_rate:.1%}")

        vig_subset["subgroup"] = pd.Categorical(
            vig_subset["subgroup"], categories=SUBGROUP_ORDER, ordered=True
        )
        vig_subset = vig_subset.sort_values("subgroup")
        print(
            vig_subset[["subgroup", "image_admit_rate", "baseline_admit_rate", "causal_effect"]]
            .to_string(index=False)
        )


# 6. Export to Excel
from openpyxl import Workbook
from openpyxl.styles import Font


def _write_section(ws, df, title, row):
    """Write a titled DataFrame block to a worksheet; return next available row."""
    ws.cell(row=row, column=1).value = title
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1).value = df.index.name or "group"
    for c_idx, col in enumerate(df.columns, start=2):
        ws.cell(row=row, column=c_idx).value = col
    row += 1
    for idx_val, data_row in df.iterrows():
        ws.cell(row=row, column=1).value = idx_val
        for c_idx, val in enumerate(data_row, start=2):
            ws.cell(row=row, column=c_idx).value = float(val) if pd.notna(val) else None
        row += 1
    return row + 1


excel_path = os.path.join(RESULTS_DIR, "analysis_results.xlsx")
wb = Workbook()
wb.remove(wb.active)

# Overall sheet
ws_overall = wb.create_sheet("Overall")
row = 1
for prompt_type in ["prompt_baseline", "prompt_ignore", "prompt_acknowledge"]:
    if prompt_type not in overall_data:
        continue
    ws_overall.cell(row=row, column=1).value = f"{prompt_type.upper()} — ALL VIGNETTES"
    ws_overall.cell(row=row, column=1).font = Font(bold=True)
    row += 2
    row = _write_section(ws_overall, overall_data[prompt_type]["subgroup"], "By Subgroup", row)
    row = _write_section(ws_overall, overall_data[prompt_type]["race"], "By Race", row)
    row = _write_section(ws_overall, overall_data[prompt_type]["gender"], "By Gender", row)
    row += 1

# Per-vignette sheets
for vignette_id in sorted(vignette_data.keys()):
    ws = wb.create_sheet(f"Vignette_{vignette_id}")
    row = 1
    for prompt_type in ["prompt_baseline", "prompt_ignore", "prompt_acknowledge"]:
        if prompt_type not in vignette_data[vignette_id]:
            continue
        d = vignette_data[vignette_id][prompt_type]
        ws.cell(row=row, column=1).value = f"{prompt_type.upper()} (baseline: {d['baseline_rate']:.1%})"
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 2
        row = _write_section(ws, d["subgroup"], "By Subgroup", row)
        row = _write_section(ws, d["race"], "By Race", row)
        row = _write_section(ws, d["gender"], "By Gender", row)
        per_case = d["per_case"].copy()
        per_case["subgroup"] = pd.Categorical(per_case["subgroup"], categories=SUBGROUP_ORDER, ordered=True)
        per_case = per_case.sort_values("subgroup").set_index("subgroup")
        row = _write_section(ws, per_case, "Per-case", row)
        row += 1

wb.save(excel_path)
print(f"\nResults saved to {excel_path}")

